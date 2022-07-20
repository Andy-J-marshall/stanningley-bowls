import openpyxl
from pathlib import Path
import json
import os
from datetime import date
import stanningleyTeamDetails

year = str(date.today().year)

stanningleyTeamNames = stanningleyTeamDetails.stanningleyTeamNames
stanningleyTeamDays = stanningleyTeamDetails.stanningleyTeamDays
stanningleyPlayers = stanningleyTeamDetails.stanningleyPlayers
duplicateTeamMemberNames = stanningleyTeamDetails.duplicateTeamMemberNames
stanningleyPlayerResults = stanningleyTeamDetails.stanningleyPlayerResults
traitorPlayers = stanningleyTeamDetails.traitorPlayers

# Spreadsheet info
homePlayerCol = 'A'
homePlayerScoreCol = 'B'
awayPlayerCol = 'C'
awayPlayerScoreCol = 'D'
homeAggCol = 'B'
awayAggCol = 'D'
homeTeamScoreCol = 'B'
awayTeamScoreCol = 'D'
homeTeamNameCol = 'A'
awayTeamNameCol = 'B'

cupText = ['qtr-finals', 'semi-finals', 'final']

# Open Excel file
path = str(Path.cwd()) + '/files/' + 'bowlsresults' + year + '.xlsx'
wb = openpyxl.load_workbook(path)

stanningleyTeamResults = []


def deduplicateNames(name):
    if name == 'Duncan Mc Phail':
        name = 'Duncan McPhail'
    if name == 'Andrew Marshall':
        name = 'Andy Marshall'
    if name == 'Stuart Watson':
        name = 'Stewart Watson'
    return name.lower()


for day in stanningleyTeamDays:
    # Goes through each sheet in turn
    sheet = wb[day]
    print('Processing ' + day)

    # Find rows in spreadsheet for Stanningley games
    startingRow = 0
    startingRowIndex = 1
    for row in sheet['A']:
        if row.value and type(row.value) is str and 'FULL RESULTS' in row.value.upper():
            startingRow = startingRowIndex
        startingRowIndex += 1

    cupGameIndex = 1
    cupGameRows = []
    for row in sheet['A']:
        if row.value and type(row.value) is str:
            if cupGameIndex > startingRow and row.value.lower() in cupText:
                for i in range(0, 11):
                    cupGameRows.append(cupGameIndex + i)
        cupGameIndex += 1

    homeIndex = 1
    homeRow = []
    for row in sheet[homeTeamNameCol]:
        if row.value and type(row.value) is str and row.value.lower() in stanningleyTeamNames:
            if homeIndex > startingRow:
                homeRow.append(homeIndex)
        homeIndex = homeIndex + 1

    awayIndex = 1
    awayRow = []
    for row in sheet[awayTeamNameCol]:
        if row.value and type(row.value) is str and row.value.lower() in stanningleyTeamNames:
            if awayIndex > startingRow:
                awayRow.append(awayIndex)
        awayIndex = awayIndex + 1

    # Find league position for Stanningley teams
    leaguePositionIndex = 1
    leaguePositionRow = 0
    currentLeaguePoints = -1
    currentLeaguePoints = -1
    leaguePositionCol = 'A'
    satLeaguePositionCol = 'B'
    leaguePointsCol = 'I'
    satLeaguePointsCol = 'A'
    monLeaguePointsCol = 'C'
    leagueTeamNameCol = 'B'
    satLeagueTeamNameCol = 'C'

    if day == 'Saturday':
        leagueTeamNameCol = satLeagueTeamNameCol
        leaguePositionCol = satLeaguePositionCol
        leaguePointsCol = satLeaguePointsCol
    if day == 'Monday':
        leaguePointsCol = monLeaguePointsCol

    for row in sheet[leagueTeamNameCol]:
        if row.value and type(row.value) is str and row.value.lower() in stanningleyTeamNames:
            leaguePosition = sheet[leaguePositionCol +
                                   str(leaguePositionIndex)].value
            leaguePoints = sheet[leaguePointsCol +
                                 str(leaguePositionIndex)].value
            if type(leaguePosition) is int:
                leaguePositionRow = leaguePositionIndex
                currentLeaguePosition = leaguePosition
            if type(leaguePoints) is int:
                currentLeaguePoints = leaguePoints
        leaguePositionIndex = leaguePositionIndex + 1
    # Find team results and scores
    awayWins = 0
    awayLosses = 0
    homeWins = 0
    homeLosses = 0
    homeDraws = 0
    awayDraws = 0
    cupWins = 0
    cupLosses = 0
    stanningleyAgg = 0
    stanningleyTotalPoints = 0
    opponentAgg = 0
    opponentTotalPoints = 0
    beaten = []
    beatenBy = []
    drawnWith = []
    results = []

    for row in range(1, sheet.max_row + 1):
        # Check if cup game
        cupGame = False
        cupCell = ''
        # Cup games are based on aggregate, not score, and are played on neutral greens
        if row > 2:
            cupCell = sheet[homeTeamNameCol + str(row - 1)].value
        if (cupCell and type(cupCell) is str) and cupCell.lower() in cupText:
            cupGame = True
            homeScore = sheet[homeTeamScoreCol + str(row + 9)].value
            awayScore = sheet[awayTeamScoreCol + str(row + 9)].value
        else:
            homeScore = sheet[homeTeamScoreCol + str(row + 10)].value
            awayScore = sheet[awayTeamScoreCol + str(row + 10)].value

        gameDate = ''
        if (row in homeRow or row in awayRow) and row > startingRow:
            gameDate = sheet[awayTeamNameCol + str(row - 1)].value
            if 'On ' in gameDate or '(from ' in gameDate:
                gameDate = sheet[awayTeamNameCol + str(row - 2)].value

        # Home games
        if row in homeRow:
            if row != leaguePositionRow:
                opponent = sheet[awayTeamNameCol + str(row)].value
                result = 'Stanningley ' + \
                    str(homeScore) + ' - ' + str(awayScore) + \
                    ' ' + opponent + ' (' + gameDate + ')'
                results.append(result)
                if homeScore > awayScore:
                    if cupGame:
                        beaten.append(opponent + ' (cup)')
                        cupWins = cupWins + 1
                    else:
                        beaten.append(opponent + ' (home)')
                        homeWins = homeWins + 1
                if homeScore < awayScore:
                    if cupGame:
                        beatenBy.append(opponent + ' (cup)')
                        cupLosses = cupLosses + 1
                    else:
                        beatenBy.append(opponent + ' (home)')
                        homeLosses = homeLosses + 1
                if awayScore == homeScore:
                    drawnWith.append(opponent + (' (home)'))
                    homeDraws = homeDraws + 1
                stanningleyAgg = stanningleyAgg + \
                    sheet[homeAggCol + str(row + 9)].value
                opponentAgg = opponentAgg + \
                    sheet[awayAggCol + str(row + 9)].value
                if not cupGame:
                    stanningleyTotalPoints = stanningleyTotalPoints + homeScore
                    opponentTotalPoints = opponentTotalPoints + awayScore

        # Away games
        if row in awayRow:
            if row != leaguePositionRow:
                opponent = sheet[homeTeamNameCol + str(row)].value
                result = opponent + ' ' + \
                    str(homeScore) + ' - ' + str(awayScore) + \
                    ' Stanningley' + ' (' + gameDate + ')'
                results.append(result)
                if awayScore > homeScore:
                    if cupGame:
                        beaten.append(opponent + ' (cup)')
                        cupWins = cupWins + 1
                    else:
                        beaten.append(opponent + ' (away)')
                        awayWins = awayWins + 1
                if awayScore < homeScore:
                    if cupGame:
                        beatenBy.append(opponent + ' (cup)')
                        cupLosses = cupLosses + 1
                    else:
                        beatenBy.append(opponent + ' (away)')
                        awayLosses = awayLosses + 1
                if awayScore == homeScore:
                    drawnWith.append(opponent + ' (away)')
                    awayDraws = awayDraws + 1
                opponentAgg = opponentAgg + \
                    sheet[homeAggCol + str(row + 9)].value
                stanningleyAgg = stanningleyAgg + \
                    sheet[awayAggCol + str(row + 9)].value
                if not cupGame:
                    stanningleyTotalPoints = stanningleyTotalPoints + awayScore
                    opponentTotalPoints = opponentTotalPoints + homeScore

    # Store team result data
    teamResults = {
        'day': day,
        'awayWins': awayWins,
        'homeWins': homeWins,
        'awayLosses': awayLosses,
        'homeLosses': homeLosses,
        'homeDraws': homeDraws,
        'awayDraws': awayDraws,
        'cupWins': cupWins,
        'cupLosses': cupLosses,
        'stanningleyAgg': stanningleyAgg,
        'stanningleyTotalPoints': stanningleyTotalPoints,
        'opponentAgg': opponentAgg,
        'opponentTotalPoints': opponentTotalPoints,
        'beaten': beaten,
        'beatenBy': beatenBy,
        'drawnWith': drawnWith,
        'leaguePosition': currentLeaguePosition,
        'leaguePoints': currentLeaguePoints,
        'results': results
    }
    stanningleyTeamResults.append(teamResults)

    # Find rows in spreadsheet for Stanningley players' games
    homePlayerIndex = 1
    homePlayerRow = []
    for homePlayer in sheet[homePlayerCol]:
        homePlayerName = homePlayer.value
        if (homePlayerName and type(homePlayerName) is str) and (homePlayerName.lower() in stanningleyPlayers or homePlayerName.lower() in duplicateTeamMemberNames):
            if homePlayerName.lower() not in traitorPlayers[day]:
                homePlayerRow.append(homePlayerIndex)
        homePlayerIndex = homePlayerIndex + 1

    awayPlayerIndex = 1
    awayPlayerRow = []
    for awayPlayer in sheet[awayPlayerCol]:
        awayPlayerName = awayPlayer.value
        if (awayPlayerName and type(awayPlayerName) is str) and (awayPlayerName.lower() in stanningleyPlayers or awayPlayerName.lower() in duplicateTeamMemberNames):
            if awayPlayerName.lower() not in traitorPlayers[day]:
                awayPlayerRow.append(awayPlayerIndex)
        awayPlayerIndex = awayPlayerIndex + 1

    # Converts points to team points
    def calculateGamePoints(points):
        if points == 21:
            gamePoints = 5
        if points < 5:
            gamePoints = 0
        if points >= 5 and points < 10:
            gamePoints = 1
        if points >= 10 and points < 15:
            gamePoints = 2
        if points >= 15 and points < 18:
            gamePoints = 3
        if points >= 18 and points < 21:
            gamePoints = 4
        return gamePoints

    # Find each players' results
    for row in range(1, sheet.max_row + 1):
        # reset variable values
        aggregate = 0
        opponentAggregate = 0
        points = 0
        opponentPoints = 0
        secondOpponent = ''
        playerName = ''
        opponentsName = ''
        pairsGame = False
        pairsPartner = ''
        opponentTeam = ''
        updateStats = False
        homeGame = None
        awayGame = None
        cupGame = False

        # Find columns
        if row in cupGameRows:
            cupGame = True

        if row in homePlayerRow:
            updateStats = True
            if not cupGame:
                homeGame = True
            stanPlayerNameCol = homePlayerCol
            stanPlayerScoreCol = homePlayerScoreCol
            opponentPlayerNameCol = awayPlayerCol
            opponentPlayerScoreCol = awayPlayerScoreCol
        if row in awayPlayerRow:
            updateStats = True
            if not cupGame:
                awayGame = True
            stanPlayerNameCol = awayPlayerCol
            stanPlayerScoreCol = awayPlayerScoreCol
            opponentPlayerNameCol = homePlayerCol
            opponentPlayerScoreCol = homePlayerScoreCol

        # Find result details
        if updateStats:
            opponentsName = sheet[opponentPlayerNameCol + str(row)].value

            if opponentsName.lower() != '*walkover*':
                playerName = sheet[stanPlayerNameCol + str(row)].value
                aggregate = sheet[stanPlayerScoreCol + str(row)].value
                opponentAggregate = sheet[opponentPlayerScoreCol +
                                          str(row)].value
                pairsGame = False
                if aggregate is None:
                    pairsGame = True
                    pairsPartner = sheet[stanPlayerNameCol +
                                         str(row - 1)].value
                    secondOpponent = sheet[opponentPlayerNameCol +
                                           str(row - 1)].value
                    aggregate = sheet[stanPlayerScoreCol + str(row - 1)].value
                    opponentAggregate = sheet[opponentPlayerScoreCol +
                                              str(row - 1)].value
                else:
                    pointsRowBelow = sheet[stanPlayerScoreCol +
                                           str(row + 1)].value
                    if pointsRowBelow is None:
                        pairsGame = True
                        pairsPartner = sheet[stanPlayerNameCol +
                                             str(row + 1)].value
                        secondOpponent = sheet[opponentPlayerNameCol +
                                               str(row + 1)].value

                playerName = deduplicateNames(playerName)

                for i in range(1, 10):
                    opponentTeamRow = sheet[awayTeamNameCol + str(row - i)]
                    if opponentTeamRow.row in homeRow:
                        opponentTeam = sheet[awayTeamNameCol +
                                             str(row - i)].value
                    if opponentTeamRow.row in awayRow:
                        opponentTeam = sheet[homeTeamNameCol +
                                             str(row - i)].value

            # Store player stats
                playerNameForResult = playerName
                if pairsGame:
                    stanningleyPlayerResults[playerName]['pairsPartners'].append(
                        pairsPartner)
                    playerNameForResult = playerName + ' & ' + pairsPartner
                    opponentsName = opponentsName + ' & ' + secondOpponent
                    stanningleyPlayerResults[playerName]['totalPairsAgg'] += aggregate
                    stanningleyPlayerResults[playerName]['totalPairsAggAgainst'] += opponentAggregate

                stanningleyPlayerResults[playerName][day.lower()]['games'] += 1

                playersResult = playerNameForResult + ' ' + \
                    str(aggregate) + ' - ' + \
                    str(opponentAggregate) + ' ' + opponentsName
                stanningleyPlayerResults[playerName]['results'].append(
                    playersResult)

                # Wins
                if aggregate > opponentAggregate:
                    stanningleyPlayerResults[playerName][day.lower(
                    )]['wins'] += 1
                    stanningleyPlayerResults[playerName]['beatenOpponents'].append(
                        opponentsName)
                    stanningleyPlayerResults[playerName]['beatenTeam'].append(
                        opponentTeam + ' (' + day + ')')
                    if pairsGame:
                        stanningleyPlayerResults[playerName]['winningPairsPartners'].append(
                            pairsPartner)
                        stanningleyPlayerResults[playerName]['pairWins'] += 1
                    if homeGame:
                        stanningleyPlayerResults[playerName]['homeWins'] += 1
                    if awayGame:
                        stanningleyPlayerResults[playerName]['awayWins'] += 1
                    if cupGame:
                        stanningleyPlayerResults[playerName]['cupWins'] += 1
                # Losses
                else:
                    stanningleyPlayerResults[playerName]['beatenBy'].append(
                        opponentsName)
                    stanningleyPlayerResults[playerName]['beatenByTeam'].append(
                        opponentTeam + ' (' + day + ')')
                    if pairsGame:
                        stanningleyPlayerResults[playerName]['losingPairsPartners'].append(
                            pairsPartner)
                        stanningleyPlayerResults[playerName]['pairLosses'] += 1
                    if homeGame:
                        stanningleyPlayerResults[playerName]['homeLosses'] += 1
                    if awayGame:
                        stanningleyPlayerResults[playerName]['awayLosses'] += 1
                    if cupGame:
                        stanningleyPlayerResults[playerName]['cupLosses'] += 1
                points = 0
                opponentPoints = 0
                if not cupGame:
                    points = calculateGamePoints(aggregate)
                    opponentPoints = calculateGamePoints(opponentAggregate)

                # Averages
                stanningleyPlayerResults[playerName]['totalAgg'] += aggregate
                stanningleyPlayerResults[playerName]['totalAggAgainst'] += opponentAggregate
                stanningleyPlayerResults[playerName][day.lower()]['aggDiff'] += aggregate - \
                    opponentAggregate
                if homeGame:
                    stanningleyPlayerResults[playerName]['totalHomeAgg'] += aggregate
                    stanningleyPlayerResults[playerName]['totalHomeAggAgainst'] += opponentAggregate
                    stanningleyPlayerResults[playerName]['totalHomePoints'] += points
                    stanningleyPlayerResults[playerName]['totalHomePointsAgainst'] += opponentPoints
                if awayGame:
                    stanningleyPlayerResults[playerName]['totalAwayAgg'] += aggregate
                    stanningleyPlayerResults[playerName]['totalAwayAggAgainst'] += opponentAggregate
                    stanningleyPlayerResults[playerName]['totalAwayPoints'] += points
                    stanningleyPlayerResults[playerName]['totalAwayPointsAgainst'] += opponentPoints
                stanningleyPlayerResults[playerName]['totalPoints'] += points
                stanningleyPlayerResults[playerName]['totalPointsAgainst'] += opponentPoints
                stanningleyPlayerResults[playerName]['dayPlayed'].append(day)

# Create JSON file
dataToExport = {
    'playerResults': stanningleyPlayerResults,
    'teamResults': stanningleyTeamResults,
    'lastUpdated': date.today().strftime("%d/%m/%Y")
}

filename = 'src/data/bowlsStats' + year + '.json'
if os.path.exists(filename):
    os.remove(filename)

with open(filename, 'w') as f:
    json.dump(dataToExport, f)
    print(filename + ' created')
