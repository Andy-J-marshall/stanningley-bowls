import openpyxl
from pathlib import Path
import os
from datetime import date

year = str(date.today().year)
filename = str(Path.cwd()) + '/files/' + 'bowlsresults' + year + '.xlsx'

if os.path.exists(filename):
    print('Excel file already exists: ' + filename)
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Monday Combined Leeds'
    wb.create_sheet('Tuesday Vets Leeds')
    wb.create_sheet('Tuesday Leeds')
    wb.create_sheet('Half Holiday Leeds')
    wb.create_sheet('Thursday Vets Leeds')
    wb.create_sheet('Saturday Leeds')
    wb.save(filename)
    print(filename + ' created')
